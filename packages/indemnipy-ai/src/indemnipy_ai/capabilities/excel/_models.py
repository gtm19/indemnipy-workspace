from dataclasses import dataclass, field
from functools import cached_property
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from typing_extensions import override

from indemnipy_ai.capabilities.excel._functions import (
    DateParsingOptions,
    _dataframe_from_range,
    _oletools_vba_parser,
    _openpyxl_table_parser,
)
from indemnipy_ai.capabilities.excel._parser_contract import _VbaParser
from indemnipy_ai.capabilities.excel._workbook_protocol import (
    VbaSummary,
    WorkbookProtocol,
    WorkbookSheet,
    WorkbookTable,
)


@dataclass
class _ExcelWorkbook(WorkbookProtocol):
    """Represents a workbook file.

    Attributes:
        filepath: Path to the workbook file.
        vba_summary: Optional summary of any detected VBA macros and analysis results.
    """

    filepath: Path

    _vba_parser: _VbaParser = _oletools_vba_parser
    date_parsing_options: DateParsingOptions | None = field(default=None, repr=False)

    @cached_property
    def _workbook(self):
        """Load the workbook using openpyxl."""
        return load_workbook(str(self.filepath), data_only=True)

    @cached_property
    def file_name(self) -> str:  # pyright: ignore[reportIncompatibleMethodOverride]
        """
        Return the name of the workbook file.

        Returns:
            The name of the workbook file as a string.
        """
        return self.filepath.name

    @cached_property
    @override
    def vba_summary(self) -> VbaSummary | None:  # pyright: ignore[reportIncompatibleMethodOverride]
        """Return a summary of any detected VBA macros and analysis results.

        Returns:
            A VbaSummary instance if macros are present, otherwise None.
        """
        vba_summary: VbaSummary = VbaSummary.from_file(
            self.filepath, parser=self._vba_parser
        )
        if not vba_summary.macros and not vba_summary.analysis_results:
            return None
        return vba_summary

    @override
    def add_table_from_range(
        self,
        sheet_name: str,
        range_str: str,
        table_name: str,
    ) -> None:
        """Add a new table to the workbook.

        Args:
            sheet_name: The name of the sheet containing the new table.
            range_str: The range string for the new table (e.g., "A1:C3").
            table_name: The name of the new table.
        """
        target_sheet = next(sheet for sheet in self.sheets if sheet.name == sheet_name)
        new_table = WorkbookTable(
            name=table_name,
            sheet_name=sheet_name,
            range=range_str,
            dataframe=_dataframe_from_range(
                self._workbook[sheet_name],
                range_str,
                date_parsing_options=self.date_parsing_options,
            ),
        )
        target_sheet.tables.append(new_table)

    @cached_property
    @override
    def sheets(self) -> list[WorkbookSheet]:  # pyright: ignore[reportIncompatibleMethodOverride]
        """Return a list of sheet names in the workbook.

        Returns:
            A list of WorkbookSheet instances representing the sheets in the workbook.
        """
        return self._get_sheets()

    def _get_sheets(self) -> list[WorkbookSheet]:
        tables = _openpyxl_table_parser(
            self._workbook, date_parsing_options=self.date_parsing_options
        )
        return [
            WorkbookSheet(
                name=sheet.title,
                range=sheet.dimensions,
                freeze_panes=sheet.freeze_panes,
                min_column=sheet.min_column,
                min_row=sheet.min_row,
                max_column=sheet.max_column,
                max_row=sheet.max_row,
                state=sheet.sheet_state,
                tables=[
                    WorkbookTable(
                        name=table.name,
                        sheet_name=table.sheet_name,
                        range=table.range,
                        dataframe=table.dataframe,
                    )
                    for table in tables.values()
                    if table.sheet_name == sheet.title
                ],
            )
            for sheet in self._workbook.worksheets
        ]

    @override
    def get_range(self, sheet_name: str, range_str: str = "A1:J100") -> list[list[Any]]:
        """
        Get the values in a specified range from a given sheet.

        Args:
            sheet_name: The name of the sheet to retrieve the range from.
            range_str: The range string (e.g., "A1:C3").

        Returns:
            A list of lists containing the values in the specified range.
        """
        sheet = self._workbook[sheet_name]
        return [[cell.value for cell in row] for row in sheet[range_str]]

    @override
    def agent_summary(self) -> str:
        """Generate a summary of the workbook for agent use.

        Returns:
            A string summary of the workbook, including sheets, tables, and VBA macros.
        """
        summary = f"Workbook: {self.file_name!r} ({self.filepath!r})\n"
        for sheet in self.sheets:
            summary += f"  Sheet: {sheet.name!r}, Range: {sheet.range!r}\n"
            for table in sheet.tables:
                summary += f"    Table: {table.name!r}, Range: {table.range!r}\n"

        return summary
